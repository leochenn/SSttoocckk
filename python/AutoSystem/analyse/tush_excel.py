
import datetime

import matplotlib.pyplot as plt
from openpyxl.reader.excel import load_workbook
from matplotlib.lines import Line2D
import matplotlib.ticker as ticker

# 该类用途：测试绘图

if __name__ == '__main__':
    wb = load_workbook(r'20211213-123078-飞凯\20211213-184146_123078.xlsx')
    ws = wb[wb.sheetnames[0]]

    # 第2行第2列
    value = ws.cell(2, 2).value
    print("第2行第2列{0}".format(value))
    print("最大行{0}".format(ws.max_row))
    print("最大列{0}".format(ws.max_column))

    x_data = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, 2).value
        x_data.append(value)

    y_data = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row, 3).value
        y_data.append(value)


    fig, ax = plt.subplots(1, 1)
    plt.plot(x_data, y_data, "red", linewidth=1)
    plt.title('title')

    custom_lines = [Line2D([0], [0], color='b', lw=4)]
    plt.legend(custom_lines, ['Hot'])

    ax.xaxis.set_major_locator(ticker.MultipleLocator(base=3))
    # ticker.MultipleLocator中的参数base是x轴显示的间隔
    plt.xlim(91512, 150059)
    plt.xticks(rotation=30)

    plt.xlabel('x')
    plt.ylabel('y')
    plt.show()

    # plt.savefig('Lekima.png', dpi=500)


