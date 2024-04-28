#!/usr/bin/python
# -*- coding: utf-8 -*-

# coding:utf-8
import os
import time

import win32con
from openpyxl import load_workbook
import chlog
import win32com.client as win32
from datetime import datetime
import win32gui

# 自动挂单完成后，手动导出撤单列表
# 读取撤单列表后输出
# 对比自动挂单后的输出，是否匹配
def manualExportListAndCheck():
    # 获取当前活动窗口句柄
    # 获取当前日期和时间
    current_date = datetime.now()
    # 格式化为"YYYYMMDD"格式的字符串
    formatted_date = current_date.strftime('%Y%m%d')

    xlsfile = r'C:\Users\Administrator\Desktop\{0}撤单查询.xls'.format(formatted_date)
    xlsxfile = xlsfile[:-3].replace('[', '').replace(']', '') + 'xlsx'
    chlog.d("xlsfile  path", xlsfile)
    chlog.d("xlsxfile path", xlsxfile)

    if not os.path.exists(xlsfile):
        chlog.e('文件不存在', xlsfile)
        exit()

    if os.path.exists(xlsxfile):
        os.remove(xlsxfile)
        chlog.e('删除已存在的文件', xlsxfile)
    if os.path.exists(xlsxfile):
        chlog.e('删除已存在的文件 失败', xlsxfile)
        exit()

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(xlsfile)
    #  xlsx: FileFormat=51
    #  xls:  FileFormat=56
    wb.SaveAs(xlsxfile, FileFormat=51)
    wb.Close()
    excel.Application.Quit()

    time.sleep(1)

    wb = load_workbook(xlsxfile)
    sheetname = wb.sheetnames[0]
    ws = wb[sheetname]
    max_row = ws.max_row + 1
    chlog.e('开始读取撤单列表', max_row)

    data = {}

    for row in range(2, max_row):
        zzName = ws.cell(row, 4).value
        zzName = zzName.replace('"', '').replace("'", "").replace('=', '')
        zzPrice = ws.cell(row, 6).value
        zzPrice = str(zzPrice)
        zzCount = ws.cell(row, 7).value
        zzCount = str(zzCount)
        # chlog.e(row, '行', zzName, zzPrice, zzCount)
        if zzName not in data:
            data[zzName] = {}

        data[zzName][zzPrice] = zzCount

    chlog.d(data)

if __name__ == '__main__':
    chlog.setTag("chtag")
    manualExportListAndCheck()