import os
from datetime import datetime
import sys
import time

import win32con
import win32gui

import log
import chlog
import math
import keyboard
import pyautogui
from openpyxl.reader.excel import load_workbook

from cancel_window import CancelWindow
from window_widget import WindowWidget
from threading import Event
import win32com.client as win32

# 深市 单位为张，一张等于100元
# 123（创业板）
# 127（主板）
# 128（中小板）

# 读取持仓excel 自动挂单卖出
# 沪市 单位为手，一手等于1000元
# 110（沪市主板）
# 111（沪市主板）
# 113（沪市主板）
# 118（沪市科创板）

# 每只转债持仓为30只
level1_count = 6
level2_count = 8
level3_count = 12

level1_price = 1.0289
level2_price = 1.0388
level3_price = 1.0488

allSellSuccess = 1
windowWidget = None

exit_flag = Event()

successSellData = {}

# 监听所有按键,按q退出程序
def on_press(key):
    if key.name in ['up', 'down', 'left', 'right']:
        chlog.e('按方向键退出执行', key.name)
        exit_flag.set()  # 设置退出标志

def doSell(name, buyCode, buyCount, buyPrice):
    if exit_flag.is_set() :
        chlog.e('执行退出1')
        sys.exit(0)

    buyCode = str(buyCode)
    buyCount = str(buyCount)
    buyPrice = str(math.floor(buyPrice * 10) / 10)
    if buyPrice.endswith('.0'):
        chlog.e('价格去掉.0', buyPrice)
        buyPrice = buyPrice.replace('.0', '')
        chlog.e('价格去掉.0', buyPrice)

    chlog.d('卖出', name, buyCode, buyPrice, buyCount)
    time.sleep(0.2)

    if not windowWidget.checkLocked():
        shelter = windowWidget.closeTipDlgBeforeOperator()
        if shelter:
            chlog.e('closeTipDlgBeforeOperator')
            time.sleep(1)

    if not windowWidget.isTabVisible(windowWidget.sellTabHwnd):
        chlog.e('卖出窗口未聚焦')
        WindowWidget.clickBtn(windowWidget.menuBar, 58, 10)
        time.sleep(0.5)

    if windowWidget.isTabVisible(windowWidget.sellTabHwnd):
        WindowWidget.focusEditCodeAndClear(windowWidget.sellTabCode)
        WindowWidget.setEditText(windowWidget.sellTabCount, buyCount)
        readSellCount = WindowWidget.getEditContent(windowWidget.sellTabCount)
        if buyCount != readSellCount:
            chlog.e("输入数量：" + str(buyCount), '读取数量', readSellCount)
        pyautogui.typewrite(buyCode)
    else:
        chlog.e('卖出窗口还是未聚焦')
        exit(0)

    time.sleep(0.2)

    WindowWidget.setEditText(windowWidget.sellTabPrice, buyPrice)
    readSellPrice = WindowWidget.getEditContent(windowWidget.sellTabPrice)
    if buyPrice != readSellPrice:
        chlog.e('输入价格', buyPrice, '读取价格', readSellPrice)

    time.sleep(2)

    if exit_flag.is_set():
        chlog.e('执行退出2')
        sys.exit(0)

    success = None

    for retry in range(2):
        # 点击买入下单
        # 注意：不能在这里使用clickBtn2
        WindowWidget.clickBtn(windowWidget.sellTabSellBtn)
        time.sleep(0.2)

        tipDialogHandle = windowWidget.getTipDlgHwndAfterCommit('卖出交易确认')
        if tipDialogHandle:
            msg = WindowWidget.getTipDlgContent(tipDialogHandle)
            if msg:
                if '股票代码' in msg:
                    # buyOrder = re.findall("\d+", msg)[0]
                    # chlog.d("委托提交成功，点击关闭", buyOrder)
                    btnHwnd = WindowWidget.getTipDlgBtn(tipDialogHandle)
                    time.sleep(0.2)
                    WindowWidget.clickBtn2(btnHwnd)
                    time.sleep(0.2)

                    # 点击后的弹窗
                    tipDialogHandle2 = windowWidget.getTipDlgHwndAfterCommit()
                    if tipDialogHandle2:
                        msg2 = WindowWidget.getTipDlgContent(tipDialogHandle2)
                        if '合同号' not in msg2:
                            chlog.e("本次交易操作失败\n", msg2)

                        btnHwnd2 = WindowWidget.getTipDlgBtn(tipDialogHandle2)
                        time.sleep(0.2)
                        WindowWidget.clickBtn2(btnHwnd2)
                        success = 1
                        if name not in successSellData:
                            successSellData[name] = {}
                        successSellData[name][buyPrice] = buyCount
                        time.sleep(0.2)
                    else:
                        chlog.e('获取委托后的弹窗句柄失败!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')

                    # 提交失败

                    # 检查委托状态， 通过读取撤单列表行数来进行判断
                    # cancelWindow = CancelWindow(windowWidget)
                    # success = cancelWindow.getListCount()
                    # if success:
                    #     log.d("卖出成功", str(datetime.datetime.now() - curr_time))
                else:
                    chlog.e('委托失败', msg)
            else:
                chlog.e('异常：委托对话框内容为空')
            break
        else:
            if retry == 2:
                raise Exception('异常：未弹出委托对话框')
            chlog.e('异常：未弹出委托对话框, 重试下单!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!')

    if not success:
        global allSellSuccess
        allSellSuccess = 0
    chlog.d("--------------------------------")

def isSh(code):
    return code.startswith('110') or code.startswith('111') or code.startswith('113') or code.startswith('118')

def sellByCodeAndCount(name, code, count, price):
    rate = 10
    if (isSh(code)):
        rate = 1

    if count < level1_count * rate:
        # 1-5
        doSell(name, code, count, price * level1_price)
    else:
        # 26-30
        if count + 1 > (level1_count * rate + level2_count * rate + level3_count * rate):
            # 卖出6
            doSell(name, code, level1_count * rate, price * level1_price)
            # 卖出8
            doSell(name, code, level2_count * rate, price * level2_price)
            # 卖出12
            doSell(name, code, level3_count * rate, price * level3_price)
        elif count + 1 > (level1_count * rate + level2_count * rate + level3_count * rate) / 2:
            # 13-25
            doSell(name, code, math.floor(count * 0.25 / rate) * rate, price * level1_price)
            doSell(name, code, math.floor(count * 0.25 / rate) * rate, price * level2_price)
            doSell(name, code, math.floor(count * 0.5 / rate) * rate, price * level3_price)
        else:
            # 12-6
            doSell(name, code, math.floor(count * 0.5 / rate) * rate, price * level1_price)
            doSell(name, code, math.floor(count * 0.5 / rate) * rate, price * level3_price)


def compareData(successSellData1):
    # 获取当前日期和时间
    current_date = datetime.now()
    # 格式化为"YYYYMMDD"格式的字符串
    formatted_date = current_date.strftime('%Y%m%d')

    xlsfile = r'C:\Users\Administrator\Documents\{0} 撤单查询.xls'.format(formatted_date)
    xlsxfile = xlsfile[:-3].replace('[', '').replace(']', '') + 'xlsx'
    chlog.d("xlsfile  path", xlsfile)
    chlog.d("xlsxfile path", xlsxfile)

    if not os.path.exists(xlsfile):
        chlog.e('文件不存在', xlsfile)
        exit()

    if os.path.exists(xlsxfile):
        os.remove(xlsxfile)
        chlog.e('删除已存在的文件', xlsxfile)
    if os.path.exists(xlsxfile):
        chlog.e('删除已存在的文件 失败', xlsxfile)
        exit()

    excel = win32.gencache.EnsureDispatch('Excel.Application')
    wb = excel.Workbooks.Open(xlsfile)
    #  xlsx: FileFormat=51
    #  xls:  FileFormat=56
    wb.SaveAs(xlsxfile, FileFormat=51)
    wb.Close()
    excel.Application.Quit()

    time.sleep(1)

    wb = load_workbook(xlsxfile)
    sheetname = wb.sheetnames[0]
    ws = wb[sheetname]
    max_row = ws.max_row + 1
    chlog.e('开始读取撤单列表', max_row)

    data = {}

    for row in range(2, max_row):
        zzName = ws.cell(row, 4).value
        zzName = zzName.replace('"', '').replace("'", "").replace('=', '')
        zzPrice = ws.cell(row, 6).value
        zzPrice = str(zzPrice)
        zzCount = ws.cell(row, 7).value
        zzCount = str(zzCount)
        # chlog.e(row, '行', zzName, zzPrice, zzCount)
        if zzName not in data:
            data[zzName] = {}

        data[zzName][zzPrice] = zzCount

    chlog.d(data)

    if data == successSellData1:
        chlog.d('结果匹配')
    else:
        chlog.e('结果不匹配')
        # 遍历并打印出不同的项目
        chlog.e('对比1')
        for outer_key, inner_dict1 in data.items():
            if outer_key not in successSellData1:
                chlog.e('在外层键中发现差异', str(outer_key), '只存在于第一个data中')
            else:
                inner_dict2 = successSellData1[outer_key]
                for inner_key, value1 in inner_dict1.items():
                    if inner_key not in inner_dict2 or inner_dict2[inner_key] != value1:
                        chlog.e('在内层键值对中发现差异', str(outer_key), str(inner_key), '的值在两个字典中不同', str(value1), 'vs',
                                str(inner_dict2.get(inner_key)))

        chlog.e('对比2')
        for outer_key, inner_dict1 in successSellData1.items():
            if outer_key not in data:
                chlog.e('在外层键中发现差异', str(outer_key), '只存在于第一个data中')
            else:
                inner_dict2 = data[outer_key]
                for inner_key, value1 in inner_dict1.items():
                    if inner_key not in inner_dict2 or inner_dict2[inner_key] != value1:
                        chlog.e('在内层键值对中发现差异', str(outer_key), str(inner_key), '的值在两个字典中不同', str(value1), 'vs',
                                str(inner_dict2.get(inner_key)))

# 导出挂单列表
def exportList():
    if not windowWidget.isTabVisible(windowWidget.cancelTab):
        chlog.d('撤单窗口未聚焦')
        WindowWidget.clickBtn(windowWidget.menuBar, 96, 10)
        time.sleep(0.5)

    if not windowWidget.isTabVisible(windowWidget.cancelTab):
        chlog.d('撤单窗口还是未聚焦')

    cancelWindow = CancelWindow(windowWidget)

    count = cancelWindow.getListCount()
    if count:
        cancelWindow.checkBuyOrderState('')

def closeWps():
    current_date = datetime.now()
    formatted_date = current_date.strftime('%Y%m%d')
    xlsfile = '{0} 撤单查询.xls - WPS Office'.format(formatted_date)
    target_window = win32gui.FindWindow(None, xlsfile)  # 根据窗口标题或类名找到目标窗口的句柄
    chlog.d('target_window', xlsfile, target_window)
    if target_window:
        chlog.e('关闭wps表格')
        win32gui.PostMessage(target_window, win32con.WM_CLOSE, 0, 0)

if __name__ == '__main__':
    log.setTag("xxx")
    chlog.setTag("xxx")

    keyboard.on_press(on_press)

    windowWidget = WindowWidget()
    windowWidget.bringToFront()

    wb = load_workbook(r'C:\Users\Administrator\Desktop\可转债轮动\收盘后的数据整理.xlsx')
    sheetname = wb.sheetnames[0]
    ws = wb[sheetname]
    max_row = ws.max_row + 1
    chlog.e('开始执行', sheetname)

    for row in range(2, max_row):
        zzName = ws.cell(row, 2).value
        if zzName is not None and len(zzName) != 0 and '转' in zzName:
            zzName = zzName.replace('"', '').replace("'", "").replace('=', '')
            zzCode = ws.cell(row, 1).value
            zzCode = ''.join([c for c in zzCode if c.isdigit()])
            zzCount = ws.cell(row, 3).value
            zzPrice = ws.cell(row, 5).value

            chlog.e("--------------------------------")
            chlog.d(row, " ", zzName, zzCode, zzPrice, zzCount)
            chlog.d("--------------------------------")
            if zzCode is None or zzCount < 1 or zzPrice < 80:
                chlog.e('错误:代码或数量不合法！！！', zzName, zzCode, zzCount, zzPrice)
            else:
                sellByCodeAndCount(zzName, zzCode, zzCount, zzPrice)
        else:
            if zzName is not None:
                chlog.e('非转债', zzName)

    if not allSellSuccess:
        chlog.e('自动挂单执行结束, 有操作失败的情况！！！')
    else:
        chlog.d(successSellData)

        xlsfile1 = r'C:\Users\Administrator\Desktop\{0}.txt'.format('successSellData')
        # 打开一个txt文件并以追加模式（不会覆盖原有内容，而是添加到文件末尾）
        with open(xlsfile1, 'a', encoding='utf-8') as f:
            # 追加字符串
            f.write(str(successSellData))

        chlog.e('自动挂单执行结束，成功! 等待手动操作，导出撤单列表')

        # 等待手动导出
        time.sleep(20)
        compareData(successSellData)

        # time.sleep(1)
        #
        # # 导出撤单列表
        # exportList()
        #
        # time.sleep(1)
        #
        # # excel窗口是否打开
        # closeWps()
        #
        # time.sleep(1)
        #
        # # 对比结果
        # compareData()
        chlog.e('完成所有操作，结束程序！')

